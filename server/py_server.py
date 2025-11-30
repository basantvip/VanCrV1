from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
import datetime

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(__file__)
CONTACT_FILE = os.path.join(BASE_DIR, 'contact.xlsx')
SHEET_NAME = 'Contacts'


def ensure_workbook(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(['Contact Number', 'Email Address', 'Message', 'DateTime', 'ActionTaken'])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(['Contact Number', 'Email Address', 'Message', 'DateTime', 'ActionTaken'])
    return wb, ws


@app.route('/save-contact', methods=['POST'])
def save_contact():
    try:
        data = request.get_json(force=True) or {}
        phone = data.get('phone', '')
        email = data.get('email', '')
        message = data.get('message', '')

        if not message and not email:
            return jsonify(ok=False, error='Missing message or email'), 400

        wb, ws = ensure_workbook(CONTACT_FILE)
        now = datetime.datetime.utcnow().isoformat()
        ws.append([phone, email, message, now, 'Pending'])
        wb.save(CONTACT_FILE)
        return jsonify(ok=True, savedTo=CONTACT_FILE)
    except Exception as e:
        app.logger.exception('save-contact failed')
        return jsonify(ok=False, error=str(e)), 500


@app.route('/', methods=['GET'])
def index():
    return 'Python contact save endpoint'


if __name__ == '__main__':
    # default port 3000 to match node server earlier
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 3000)), debug=True)
